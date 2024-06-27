import requests
import json
import re
import openpyxl

from openpyxl import Workbook, load_workbook
from tkinter import filedialog, ttk
from googletrans import Translator
from datetime import datetime
import pandas as pd
import tkinter as tk
import os
# day=input("Enter current day, it should be in proper spelling and case, ex: Friday: ")
print("Ensure that `numbers.txt` exist ar same path as script")

def get_current_day():
    # Get today's day of the week
    today = datetime.now().strftime('%A')
    return today
def translate_to_japanese(text):
    translator = Translator()
    translated_text = translator.translate(text, dest='ja').text
    return translated_text
def open_file():
    file_path = filedialog.askopenfilename(
        filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx;*.xls")]
    )
    if file_path:
        print(f"Selected file: {file_path}")
        process_file(file_path)


def process_file(file_path):
    try:
        numbers = []
        total_rows = 0
        print(file_path)
        global data
        idx=0
        if file_path.endswith('.csv'):
            a=read_numbers_from_file(file_path)
            for number in a:
                number=number.replace('"', '')
                number = number.strip()
                if not number.startswith("0"):
                    number = "0" + number
                print("Started", number)

                data = scrape_details(number)
                populate_excel(data)
                if idx == 1:
                    progress_label.config(text="完了")
                else:
                    progress_label.config(text="読み込み中")
                root.update_idletasks()
                # progress_bar["value"] = idx + 1
                # progress_label.config(text=f"Processing row {idx + 1}")
                # root.update_idletasks()
                # idx+=1
        else:
            a = openpyxl.load_workbook(file_path)
            sheet = a.active
            # idx=1
            for row in sheet.iter_rows(values_only=True):
                for  cell in row:
                    if idx == 1:
                            progress_label.config(text="完了")
                    else:
                        progress_label.config(text="読み込み中")
                    if cell is not None: 
                        print(type(cell))      
                        print(cell)                 
                        if not isinstance(cell, str):
                            cell = str(cell)
                        print(type(cell))      
                        print(cell)
                        cell=cell.replace('"', '')
                        cell = cell.strip()
                        if not cell.startswith("0"):
                            cell = "0" + cell
                        # print(cell)
                        # print("Started", number)
                        # number=cell
                        print("Started", cell)
                        data = scrape_details(cell)
                        populate_excel(data)
                        
                        # progress_bar["value"] = idx + 1
                        # progress_label.config(text=f"Processing row {idx + 1}")
                        # root.update_idletasks()
                        # idx+=1
                    root.update_idletasks()
            total_rows = len(numbers)
            print(total_rows)
        idx=1
        print("idx to 1")
        
        root.update_idletasks()
        if idx == 1:
            progress_label.config(text="完了")
        else:
            progress_label.config(text="読み込み中")

        # progress_bar["maximum"] = total_rows
        # for idx, number in a:
    # print(type(number))
            
        # total_rows = len(data)
        # progress_bar["maximum"] = total_rows

        # for idx, row in data.iterrows():
        #     # Update the progress bar and label
           

        #     # Processing logic: get the phone number and run your function
        #     phone_number = str(row.iloc[0])  # Convert number to string
        #     run_custom_function(phone_number)  # Run your custom function on the phone number

    except Exception as e:
        progress_label.config(text="Error processing file")
        print(f"Error: {e}")

def populate_excel(data_dict, filename='output.xlsx'):
    if data_dict is None:
        print("No data provided to write to Excel.")
        return

    try:
        # Check if the file already exists
        if os.path.exists(filename):
            print(f"File '{filename}' exists. Loading existing workbook.")
            workbook = load_workbook(filename)
            sheet = workbook.active
        else:
            print(f"File '{filename}' does not exist. Creating new workbook.")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(list(data_dict.keys()))  # Add headers

        # Append the values as a new row
        sheet.append(list(data_dict.values()))

        # Save the workbook
        workbook.save(filename)
        print(f"Data successfully written to '{filename}'.")

    except Exception as e:
        print(f"Error while writing to Excel: {e}")

def run_custom_function(phone_number):
    # Your custom function logic goes here
    
    try:
        # print(type(phone_number))
        print("Started", phone_number)
        data = scrape_details(phone_number)
        populate_excel(data)
    except Exception as e:
        print(e)
    # Add your specific code here
    # For example, you can use an API call, data manipulation, etc.

def read_numbers_from_file(filename):
    try:
        with open(filename, 'r', encoding='utf-8') as file:
            lines = file.readlines()
    except UnicodeDecodeError:
        with open(filename, 'r', encoding='latin-1') as file:
            lines = file.readlines()

    numbers = []
    for line in lines:
        line = line.strip()
        try:
            number = line
            numbers.append(number)
        except ValueError:
            # Ignore lines that are not valid integers
            continue

    return numbers

def extract_reviews(input_string):
    try:
        # Define the regex pattern to find the number of reviews
        pattern = re.compile(r'\d+ reviews')
        
        # Search for the pattern in the input string
        match = pattern.search(input_string)
        
        # If a match is found, return it, otherwise return None
        if match:
            return match.group(0)
        else:
            return "無"
    except Exception as e:
        # If any error occurs, return an empty string
        return "無"

def extract_ratings(input_string):
    try:
        # Define the regex pattern to find numbers like 4.3 or 3.2
        pattern = re.compile(r'\b\d\.\d\b')
        
        # Find all matches in the input string
        matches = pattern.findall(input_string)
        
        # Return the first match or an empty string if no matches are found
        if matches:
            return matches[0]
        else:
            return "無"
    except Exception as e:
        # If any error occurs, return an empty string
        return "無"

def find_hex_pairs_indices(superString):
    try:
        # Define the regular expression pattern to match the specified string format
        pattern = r'\b0x[0-9a-fA-F]{16}:0x[0-9a-fA-F]{16}\b'
        
        # Use re.finditer to find all matches and their start positions
        matches = re.finditer(pattern, superString)
        
        # Extract and return the start indices of all matches
        indices = [match.start() for match in matches]
        if indices==[]:
                pattern = r'\b0x[0-9a-fA-F]{16}:0x[0-9a-fA-F]{15}\b'
                matches = re.finditer(pattern, superString)
                indices = [match.start() for match in matches]
                return indices
        
        return indices
    except Exception as e:
        # If any error occurs, return an empty list
        return []

burp0_cookies = {"OGPC": "19037049-1:", "AEC": "AQTF6HwmJ2zkfW0KGRCQKPO4OYXzdonVOeuC1K3WpngvO7xaUzrv2wlryw", "NID": "515=N63sxz0maRkitFa0zfuUzJkakTwsQWTc0UQfSRF1s2Fbm-cDkXKEKYZCmTPTxrm_IZ3bdlaoU_CLu88SvRickztT3Phr9_7cmheDHKIliPK9wYDpQfBzz8w1pWu9U0rtV5WdOSlq7iMNZwuqQ_XPN21VDZLNZzguEreVn-FtYGPE-2rRNLZXPIxhnD1uuivx9u3RaNM1BJWWpRj3UIHmuKzevEHBcNpKs757-O0iWd-SR8JM6dvw9V0FjvmvJpNbNtU_E5LN7kCKDh9UxSHxQ_w_mRA", "1P_JAR": "2024-06-21-10"}
burp0_headers = {"Sec-Ch-Ua": "\"Not/A)Brand\";v=\"8\", \"Chromium\";v=\"126\", \"Google Chrome\";v=\"126\"", "X-Maps-Diversion-Context-Bin": "CAE=", "Sec-Ch-Ua-Mobile": "?0", "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36", "Sec-Ch-Ua-Arch": "\"x86\"", "Sec-Ch-Ua-Full-Version": "\"126.0.6478.62\"", "Sec-Ch-Ua-Platform-Version": "\"10.0.0\"", "X-Goog-Ext-353267353-Jspb": "[null,null,null,147535]", "Sec-Ch-Ua-Full-Version-List": "\"Not/A)Brand\";v=\"8.0.0.0\", \"Chromium\";v=\"126.0.6478.62\", \"Google Chrome\";v=\"126.0.6478.62\"", "Sec-Ch-Ua-Bitness": "\"64\"", "Sec-Ch-Ua-Model": "\"\"", "Sec-Ch-Ua-Wow64": "?0", "Sec-Ch-Ua-Platform": "\"Windows\"", "Accept": "*/*", "Sec-Fetch-Site": "same-origin", "Sec-Fetch-Mode": "cors", "Sec-Fetch-Dest": "empty", "Referer": "https://www.google.com/", "Accept-Encoding": "gzip, deflate", "Accept-Language": "en-US,en;q=0.9,en-IN;q=0.8", "Priority": "u=1, i"}

# def scrape_details(number):
#     burp0_url = "https://www.google.com:443/search?tbm=map&authuser=0&hl=en&gl=jp&pb=!4m12!1m3!1d3243.712031834382!2d130.76993367560658!3d33.642619873313905!2m3!1f0!2f0!3f0!3m2!1i1488!2i750!4f13.1!7i20!10b1!12m17!1m2!18b1!30b1!2m3!5m1!6e2!20e3!10b1!12b1!13b1!16b1!17m1!3e1!20m3!5e2!6b1!14b1!19m4!2m3!1i360!2i120!4i8!20m57!2m2!1i203!2i100!3m2!2i4!5b1!6m6!1m2!1i86!2i86!1m2!1i408!2i240!7m42!1m3!1e1!2b0!3e3!1m3!1e2!2b1!3e2!1m3!1e2!2b0!3e3!1m3!1e8!2b0!3e3!1m3!1e10!2b0!3e3!1m3!1e10!2b1!3e2!1m3!1e9!2b1!3e2!1m3!1e10!2b0!3e3!1m3!1e10!2b1!3e2!1m3!1e10!2b0!3e4!2b1!4b1!9b0!22m6!1sD1t1ZuunO4aRseMPz_OyoAs%3A1!2s1i%3A0%2Ct%3A11886%2Cp%3AD1t1ZuunO4aRseMPz_OyoAs%3A1!7e81!12e5!17sD1t1ZuunO4aRseMPz_OyoAs%3A93!18e15!24m95!1m31!13m9!2b1!3b1!4b1!6i1!8b1!9b1!14b1!20b1!25b1!18m20!3b1!4b1!5b1!6b1!9b1!12b1!13b1!14b1!17b1!20b1!21b1!22b1!25b1!27m1!1b0!28b0!31b0!32b0!33m1!1b0!10m1!8e3!11m1!3e1!14m1!3b1!17b1!20m2!1e3!1e6!24b1!25b1!26b1!29b1!30m1!2b1!36b1!39m3!2m2!2i1!3i1!43b1!52b1!54m1!1b1!55b1!56m1!1b1!65m5!3m4!1m3!1m2!1i224!2i298!71b1!72m19!1m5!1b1!2b1!3b1!5b1!7b1!4b1!8m10!1m6!4m1!1e1!4m1!1e3!4m1!1e4!3sother_user_reviews!6m1!1e1!9b1!89b1!103b1!113b1!117b1!122m1!1b1!125b0!126b1!26m4!2m3!1i80!2i92!4i8!30m28!1m6!1m2!1i0!2i0!2m2!1i530!2i750!1m6!1m2!1i1438!2i0!2m2!1i1488!2i750!1m6!1m2!1i0!2i0!2m2!1i1488!2i20!1m6!1m2!1i0!2i730!2m2!1i1488!2i750!34m18!2b1!3b1!4b1!6b1!8m6!1b1!3b1!4b1!5b1!6b1!7b1!9b1!12b1!14b1!20b1!23b1!25b1!26b1!37m1!1e81!42b1!47m0!49m8!3b1!6m2!1b1!2b1!7m2!1e3!2b1!8b1!50m4!2e2!3m2!1b1!3b1!61b1!67m2!7b1!10b1!69i696&q="+number+"&oq="+number+"&gs_l=maps.3..115i144k1.36531.36531.2.39517.1.1.....147.147.0j1.1.....0....1..maps..0.1.148.0..38i39k1j38.&tch=1&ech=2&psi=D1t1ZuunO4aRseMPz_OyoAs.1718967055647.1"
#     try:
#         b=requests.get(burp0_url, headers=burp0_headers, cookies=burp0_cookies)
#         a=b.text
#         z=a[:-6]
#         s=json.loads(z)
#         b=s["d"]
#         #r is main string
#         r=b[0:b.find("https://lh5.googleusercontent.com")]
#         n = find_hex_pairs_indices(r)
#         name = r[n[0]+40:r.find("\"",n[0]+40)]
#         print(name)
#         address1 = r.find("null,null,null,",n[0]) + 16
#         address2 = r.find("\"",address1)

#         address = r[address1:address2]
#         print(address)


#         rating = extract_ratings(r)
#         print(rating)

#         reviews = extract_reviews(r)
#         print(reviews)

    
#         schedule1 = r.find("[[\""+day+"\"")
#         schedule2 = r.find("M\"]",schedule1+15)

#         w=r[schedule1:schedule2+3]
#         w=w.replace('null','')
#         w=w.replace('\u202f','')
#         w=w.replace('[','')
#         w=w.replace(']','')
#         schedule=w
#         schedule=schedule.replace("\""+day+"\"","")
#         print(schedule)
#         rest_day1=r.find("[\"Closed\"]")
#         rest_day2=rest_day1-9
#         if rest_day1 == -1:
#             rest_day = "NULL"
#         else:
#             rest_day=r[rest_day2:rest_day1]
#         print(rest_day1,rest_day2)
#         return {"番号": number,"名前": name,"スケジュール": schedule,"休日":rest_day, "評価": rating, "レビュー": reviews, "住所": address,}
#     except Exception as e:
#         with open("not.txt", "a") as f:
#             f.write(number+"\n")
#         print(f"An error occurred: {e}")
#         return None
def remove_day_from_schedule(schedule):
    # Regular expression to match day names (Monday to Sunday) followed by a comma and optional whitespace
    day_pattern = re.compile(r'\b(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday),\s*')
    
    # Remove day names from the schedule string
    processed_schedule = day_pattern.sub('', schedule)
    
    # Check for any remaining day name within quotes and remove it
    processed_schedule = re.sub(r'\"(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\"', '', processed_schedule)
    
    return processed_schedule
def process_schedule(schedule):
    if "Closed" in schedule:
        return "Closed"
    elif schedule.strip() == "":
        return "無"
    else:
        return schedule
def remove_reviews_word(input_string):
    cleaned_string = input_string.replace(" reviews", "")
    return cleaned_string.strip()
def scrape_details(number):
    day=today = datetime.now().strftime('%A')
    burp0_url = "https://www.google.com:443/search?tbm=map&authuser=0&hl=en&gl=jp&pb=!4m12!1m3!1d3243.712031834382!2d130.76993367560658!3d33.642619873313905!2m3!1f0!2f0!3f0!3m2!1i1488!2i750!4f13.1!7i20!10b1!12m17!1m2!18b1!30b1!2m3!5m1!6e2!20e3!10b1!12b1!13b1!16b1!17m1!3e1!20m3!5e2!6b1!14b1!19m4!2m3!1i360!2i120!4i8!20m57!2m2!1i203!2i100!3m2!2i4!5b1!6m6!1m2!1i86!2i86!1m2!1i408!2i240!7m42!1m3!1e1!2b0!3e3!1m3!1e2!2b1!3e2!1m3!1e2!2b0!3e3!1m3!1e8!2b0!3e3!1m3!1e10!2b0!3e3!1m3!1e10!2b1!3e2!1m3!1e9!2b1!3e2!1m3!1e10!2b0!3e3!1m3!1e10!2b1!3e2!1m3!1e10!2b0!3e4!2b1!4b1!9b0!22m6!1sD1t1ZuunO4aRseMPz_OyoAs%3A1!2s1i%3A0%2Ct%3A11886%2Cp%3AD1t1ZuunO4aRseMPz_OyoAs%3A1!7e81!12e5!17sD1t1ZuunO4aRseMPz_OyoAs%3A93!18e15!24m95!1m31!13m9!2b1!3b1!4b1!6i1!8b1!9b1!14b1!20b1!25b1!18m20!3b1!4b1!5b1!6b1!9b1!12b1!13b1!14b1!17b1!20b1!21b1!22b1!25b1!27m1!1b0!28b0!31b0!32b0!33m1!1b0!10m1!8e3!11m1!3e1!14m1!3b1!17b1!20m2!1e3!1e6!24b1!25b1!26b1!29b1!30m1!2b1!36b1!39m3!2m2!2i1!3i1!43b1!52b1!54m1!1b1!55b1!56m1!1b1!65m5!3m4!1m3!1m2!1i224!2i298!71b1!72m19!1m5!1b1!2b1!3b1!5b1!7b1!4b1!8m10!1m6!4m1!1e1!4m1!1e3!4m1!1e4!3sother_user_reviews!6m1!1e1!9b1!89b1!103b1!113b1!117b1!122m1!1b1!125b0!126b1!26m4!2m3!1i80!2i92!4i8!30m28!1m6!1m2!1i0!2i0!2m2!1i530!2i750!1m6!1m2!1i1438!2i0!2m2!1i1488!2i750!1m6!1m2!1i0!2i0!2m2!1i1488!2i20!1m6!1m2!1i0!2i730!2m2!1i1488!2i750!34m18!2b1!3b1!4b1!6b1!8m6!1b1!3b1!4b1!5b1!6b1!7b1!9b1!12b1!14b1!20b1!23b1!25b1!26b1!37m1!1e81!42b1!47m0!49m8!3b1!6m2!1b1!2b1!7m2!1e3!2b1!8b1!50m4!2e2!3m2!1b1!3b1!61b1!67m2!7b1!10b1!69i696&q="+number+"&oq="+number+"&gs_l=maps.3..115i144k1.36531.36531.2.39517.1.1.....147.147.0j1.1.....0....1..maps..0.1.148.0..38i39k1j38.&tch=1&ech=2&psi=D1t1ZuunO4aRseMPz_OyoAs.1718967055647.1"
    try:
        b=requests.get(burp0_url, headers=burp0_headers, cookies=burp0_cookies)
        a=b.text
        z=a[:-6]
        s=json.loads(z)
        b=s["d"]
        #r is main string
        r=b[0:b.find("https://lh5.googleusercontent.com")]
        n = find_hex_pairs_indices(r)
        name = r[n[0]+40:r.find("\"",n[0]+40)]
        print("name",name)
        address1 = r.find("null,null,null,",n[0]) + 16
        address2 = r.find("\"",address1)

        address = r[address1:address2]
        print("Address",address)


        rating = extract_ratings(r)
        print("rating",rating)

        reviews = extract_reviews(r)
        print("Reviews",reviews)

        reviews=remove_reviews_word(reviews)
        schedule1 = r.find(day) + len(day)
        schedule2 = r.find("\"]",schedule1+len(day))
        schedule = r[schedule1:schedule2]
        schedule = schedule.replace("\"","")
        schedule = schedule.replace(",","")
        schedule = schedule.replace("[","")
        print("schedule",schedule)
        if len(schedule)>15:
            schedule = "無"
        else:
            print(type(schedule))
            schedule=translate_to_japanese(schedule)
        print(type(schedule))
        rest_day1=r.find("[\"Closed\"]")
        rest_day2=rest_day1-10
        if rest_day1 == -1:
            rest_day = "記載なし"
        else:
            rest_day=r[rest_day2:rest_day1]
        rest_day = rest_day.replace("\"","")
        rest_day = rest_day.replace(",","")
        rest_day = rest_day.replace("[","")
        print("rest_day", rest_day)
        print(rest_day)
        name=translate_to_japanese(name)
        # schedule=translate_to_japanese(schedule)
        rest_day=translate_to_japanese(rest_day)
        rating=translate_to_japanese(rating)
        reviews=translate_to_japanese(reviews)
        address=translate_to_japanese(address)
        return {"番号": number,"名前": name,"スケジュール": schedule,"休日":rest_day, "評価": rating, "レビュー": reviews, "住所": address,}
    except Exception as e:
        with open("not.txt", "a") as f:
            f.write(number+"\n")
        print(f"An error occurred: {e}")
        return None

# Create the main window
root = tk.Tk()
root.title("ファイルプロセッサ")

main_frame = tk.Frame(root, padx=10, pady=10)
main_frame.pack(expand=True, fill=tk.BOTH)

open_button = tk.Button(main_frame, text="ファイルを開く", command=open_file)
open_button.pack(pady=10)

progress_label = tk.Label(main_frame, text="ファイルが選択されていません")
progress_label.pack()

root.mainloop()

# a=read_numbers_from_file("numbers.csv")

# for number in a:
#     # print(type(number))
#     print("Started", number)
#     data = scrape_details(number)
#     populate_excel(data)




    
