from datetime import datetime
import re
import sqlite3
import sys
import os
from PySide6.QtCore import QUrl
from PySide6.QtGui import *
from PySide6.QtWidgets import QApplication
from PySide6.QtQml import QQmlApplicationEngine
from PySide6.QtCore import *
from PySide6.QtWidgets import *
from PySide6.QtCore import *
from TrackStar import check_key, TrackNumber
import configparser
import json
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from TrackModel import TrackModel, HistoryModel, categorize_date, LoginModel, FormatModel
import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from sqlalchemy.pool import NullPool
import pandas as pd
# from docx import Document
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors


engine = create_engine('sqlite:///general.db', poolclass=NullPool)
Session = sessionmaker(bind=engine)


class GlogalVariable:
    track_results = list()
    pattern = r'^\d{10,13}$'
 

def col_format(maxcol):
    def generate_label(n):
        if n <= 0:
            return ""
        elif n <= 26:
            return chr(64 + n)
        else:
            quotient, remainder = divmod(n - 1, 26)
            return generate_label(quotient) + generate_label(remainder + 1)
    labels = [generate_label(i) for i in range(1, maxcol + 1)]
    return labels



def filter_dict(original_dict, keys_to_keep):
    return {key: original_dict[key] for key in keys_to_keep if key in original_dict}


    
def decode_column_label(label):
    label = label.upper()  # Convert the label to uppercase to handle lowercase input
    num = 0

    for char in label:
        num = num * 26 + (ord(char) - ord('A')) + 1
    return num



def get_max_row_and_column(sheet):
    max_col = 0
    max_row = 0
    column_due = 3
    row_due = 3

    # Iterate through rows to find the maximum row with non-empty cells
    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
        max_row += 1
        if all(cell_value is None for cell_value in row):
            row_due -= 1
            max_row -= 1
            if row_due == 0:
                break

    # Iterate through columns to find the maximum column with non-empty cells
    for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
        max_col += 1
        if all(cell_value is None for cell_value in col):
            column_due -= 1
            max_col -= 1
            if column_due == 0:
                break 

    return max_row, max_col



def get_sheet_list(sheet, self):
    max_row, max_col = get_max_row_and_column(sheet)
    cols = col_format(max_col)
    sheetList = list()
    self.highList = list()
    self.allCellValues = dict()
    
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_col=max_col, max_row=max_row), start=1):
        row_info = [
            {'value': str(index), 'width': 40, "color": "#F9F9F9", "coordinate": "x"}
        ]

        highlighted = 0  
        
        for column, cell in zip(cols, row):
            cellValue = " " if cell.value is None else str(cell.value)
            color, highlight =  ["#E3E0F6", 1] if re.match(GlogalVariable.pattern , cellValue) else ["white", 0]
            highlighted += highlight
            if highlight == 1:
                self.highList.append({"coordinate": cell.coordinate, 'value': cellValue})
            row_info.append(
                {'value': cellValue,
                 'width': sheet.column_dimensions[column].width * 7,
                 "color": color,
                 "coordinate": cell.coordinate,
                 "row": index,
                 "col": cols.index(column),
                 f"{cell.coordinate}" : cellValue} 
            )
            self.allCellValues[ f"{cell.coordinate}"] = cellValue 
        sheetList.append({"row": json.dumps(row_info) , "id": index, "highlighted": highlighted}) # remove the highlighted if not needed
    return sheetList




class Worker(QObject):
    finished = Signal()
    progress = Signal(list)
    check = Signal(dict)
    single = Signal(dict)
    multiple = Signal(dict)
    message = Signal(list)
    load = Signal(dict)
    fetch = Signal(dict)
    save = Signal(dict)


      

    def __init__(self, tracker=None, track_list=None, track_name=None, excel_name=None, work_sheet=None, start_row=None,
                 start_col=None, max_col=None, api_key=None, username=None, number=None, awbs=None, resume=0, mode=None, address=None,
                  path=None , extension=None,  headers=None):
        super(Worker, self).__init__()
        self.tracker = tracker
        self.track_list = track_list
        self.number = number
        self.api_key = api_key
        self.username=username

        
        self.awbs = awbs
        self.track_name = track_name
        self.work_sheet = work_sheet
        self.start_row = start_row
        self.start_col = start_col
        self.max_col = max_col
        self.excel_name = excel_name
        
        self.resume = resume
        self.mode = mode
        self.address = address
        self.path = path
        self.extension = extension
        self.headers = headers 
        self.stopped_flag = False 
        self.sheetList = list()
        


    def stop(self):
        self.stopped_flag = True      

    def check_login_work(self):
        """Long-running task."""
        self.check.emit(check_key(api_key=self.api_key, username=self.username )) 
        self.finished.emit()


    def update_sheet(self):
        self.sheetList = get_sheet_list(self.sheet, self)
               
        
    def load_excel_work(self):
        try:
            self.excel_name = str(self.excel_name)[8:]
            self.book = xl.load_workbook(self.excel_name)
            self.sheet = self.book.active
            max_row, max_col = get_max_row_and_column(self.sheet)
            rows = [i for i in range(1, max_row + 1)]
            cols = [{'value': " ", 'width': 40 , "color":"#F9F9F9", "coordinate": "xy" }] + [{'value': column, 'width': self.sheet.column_dimensions[column].width*7 , "color":"#F9F9F9", "coordinate": "y"  } for column in col_format(max_col)]
            loaded = { "rows": json.dumps(rows) , "cols": json.dumps(cols) , "sheet_names": json.dumps(self.book.sheetnames) , "active_sheet_index": self.book.sheetnames.index(self.sheet.title), "access": True }  #  "sheetList": json.dumps(self.sheetList)  
            self.update_sheet()
            self.load.emit(loaded)     
            print("Loaded")
        except xl.utils.exceptions.InvalidFileException:
            self.load.emit({ "access": False, "status": "File Error"})
        except Exception as e:
            print(str(e))
            self.load.emit({"access": False, "status": str(e)})
        self.finished.emit()        
    

    def track_single_work(self):
        """Long-running task."""
        self.track_single = TrackNumber(api_key=self.api_key)
        result = self.track_single.track_number(self.number)           
        # print(result)
        result.update({"SN" : 1 })        
        if result["truth"]:
          
            with Session() as session: 
                present = session.query(HistoryModel).filter_by(name=self.number).first()      
                if present is None:         
                    new_data = HistoryModel(name=self.number, mode=self.mode, address=json.dumps(result), completed=True,  progress=1,  total=1, timestamp=str(datetime.now().strftime("%A, %B %d %Y %I:%M%p")), last_updated=str(datetime.now().strftime("%A, %B %d %Y %I:%M%p")), number=f"{self.number}")
                    session.add(new_data)
                    session.commit()
                else:
                    present.address, present.last_updated = json.dumps(result),  str(datetime.now().strftime("%A, %B %d %Y %I:%M%p"))
                    session.commit()
        self.single.emit(result)
        self.finished.emit()



    def track_multiple_work(self):
        """Long-running task."""
        self.stopped_flag = False

        with Session() as session:
            current_track_history = session.query(HistoryModel).filter_by(name=self.track_name).first()     

            if current_track_history is None:           
                track_engine = create_engine(f'sqlite:///trackdb/{self.track_name}.db')
                TrackModel.metadata.create_all(track_engine)
                TrackSession = sessionmaker(bind=track_engine)
                track_session = TrackSession()
                new_data = HistoryModel(name=self.track_name, mode=self.mode, address=f"trackdb/{self.track_name}.db", completed=False,  progress=0,  total=len(self.awbs), timestamp=str(datetime.now().strftime("%A, %B %d %Y %I:%M%p")), last_updated=str(datetime.now().strftime("%A, %B %d %Y %I:%M%p")), number=",".join(map(str, self.awbs)))
                session.add(new_data)
                session.commit()
            else:
                if current_track_history.completed and len(self.awbs) == current_track_history.total:
                    self.multiple.emit({"truth": False, "status": "Completed", "path": self.track_name })
                    self.finished.emit()
                    return 
                track_engine = create_engine(f'sqlite:///trackdb/{self.track_name}.db')
                TrackSession = sessionmaker(bind=track_engine)
                track_session = TrackSession()            
                if self.resume == 0:
                    track_details = track_session.query(TrackModel).all() 
                    # Convert each row to a dictionary
                    track_dicts = [track_detail.__dict__ for track_detail in track_details]
                    self.resume = len(track_dicts)
                    for data in track_dicts:
                        self.multiple.emit(data)
 
        self.track_range = TrackNumber(api_key=self.api_key)
        # total = len(self.awbs)
        counter = 0
        for result in self.track_range.track_numbers(self.awbs[self.resume:]):
            if not self.stopped_flag:
                # print(result) 
                serial_number = self.resume + counter + 1            
                result.update({"SN" : serial_number })                
                self.multiple.emit(result)
                
                if result["truth"]:
                    track_new_row = TrackModel(**result)
                    track_session.add(track_new_row)
                    track_session.commit()            
                    
                    counter += 1 
                    with Session() as session:
                        current_track_history = session.query(HistoryModel).filter_by(name=self.track_name).first()                       
                        current_track_history.progress = serial_number  
                        current_track_history.completed = current_track_history.progress  == current_track_history.total
                        current_track_history.last_updated=str(datetime.now().strftime("%A, %B %d %Y %I:%M%p")) 
                        session.commit()                                         
            else:
                self.multiple.emit({"truth": False, "status": "Stopped" })
                # self.finished.emit()
                counter = 0
                break
                                          
        track_session.close()       
        self.finished.emit()

 
    def fetch_export_work(self):     
        if self.mode == "Single":
            with  Session() as session:
                single_history = session.query(HistoryModel).filter_by(name=self.address).first()
            data = json.loads(single_history.address)
            GlogalVariable.track_results = [data]
            self.fetch.emit(data) 
        else:
            if os.path.exists(f'trackdb/{self.address}.db'):
                print("database exists")                 
                track_engine = create_engine(f'sqlite:///trackdb/{self.address}.db')
                TrackSession = sessionmaker(bind=track_engine)
                track_session = TrackSession()                
                track_details = track_session.query(TrackModel).all() 
                # Convert each row to a dictionary
                GlogalVariable.track_results = [track_detail.__dict__ for track_detail in track_details]
                for data in GlogalVariable.track_results:
                    self.fetch.emit(data)
                    # print(data)
                track_session.close()     
        self.finished.emit()
   

 
    def fetch_export_work(self):     
        if self.mode == "Single":
            with  Session() as session:
                single_history = session.query(HistoryModel).filter_by(name=self.address).first()
            data = json.loads(single_history.address)
            GlogalVariable.track_results = [data]
            self.fetch.emit(data) 
        else:
            if os.path.exists(f'trackdb/{self.address}.db'):
                print("database exists")                 
                track_engine = create_engine(f'sqlite:///trackdb/{self.address}.db')
                TrackSession = sessionmaker(bind=track_engine)
                track_session = TrackSession()                
                track_details = track_session.query(TrackModel).all() 
                # Convert each row to a dictionary
                GlogalVariable.track_results = [track_detail.__dict__ for track_detail in track_details]
                for data in GlogalVariable.track_results:
                    self.fetch.emit(data)
                    # print(data)
                track_session.close()     
        self.finished.emit()


    def save_fetched_work(self):           
        if len(GlogalVariable.track_results) > 0:
            wb = xl.Workbook()
            ws = wb.active
         
            columns = col_format(len(self.headers))

            ws.merge_cells(f'A1:{columns[-1]}1')
            ws['A1'].value = "THIS AUTO TRACKING RESULTS IS BROUGHT TO YOU BY WEMMYNEAT TECHNOLOGIES"
            ws['A1'].font = Font(name='Calibri', size=14, bold=True, italic=True)
            ws['A1'].alignment = Alignment(horizontal="center")

            for column, [head ,width] in zip(columns, self.headers):
                ws.column_dimensions[column].width = width/5
                ws[f'{column}2'].value = head
                ws[f'{column}2'].font = Font(name='Calibri', size=12, bold=True)
      
            # Write the data from the dictionaries to the Excel sheet
            for row_num, data_dict in enumerate(GlogalVariable.track_results, start=3):
                for col_num,[ key, width] in enumerate(self.headers, start=1):
                    ws.cell(row=row_num, column=col_num, value=data_dict[key])
            if self.extension == "xlsx":
                try:
                    # Save the Excel file
                    wb.save( self.path[8:])
                    file_name = self.path[self.path.rfind('/')+1:]
                    self.save.emit({"truth": True, "status": f"{file_name} \n Sucessfuly Saved ✅"})
                except PermissionError as e:
                    self.save.emit({"truth": False, "status": str(e)})
                    print("Permission Error")
                GlogalVariable.track_results = [ ]
                print("Saved successfully")
                                    
        self.finished.emit()



        

def threader (worker, thread, main_fn, *args):
    thread.quit()
    worker.moveToThread(thread)
    # Step 5: Connect signals and slots
    thread.started.connect(main_fn)
    worker.finished.connect(thread.quit)
    worker.finished.connect(worker.deleteLater)
    thread.finished.connect(thread.deleteLater)
    for [sign, fn] in args:
        sign.connect(fn)
    # Step 6: Start the thread
    thread.start()
    return worker, thread




class MainWindow(QObject):
    check = Signal(list, arguments=['checked'])
    single = Signal(dict, arguments=['sign'])
   
    excels = Signal(dict, arguments=['excel'])
   
    pastes  = Signal(dict, arguments=['paste'])
    load = Signal(dict, arguments=['loads'])
    fetch = Signal(dict, arguments=['fetches'])
    save = Signal(dict, arguments=['save'])
    

    

    def __init__(self, parent=None):
        super().__init__(parent)
        self.tracker = None
        self._track_file = None
        self.range_worker = None
        self.range_thread = None 
        self._sheetList = list()


    @Slot(result=list)
    def history(self):
        with Session() as session:      
            history_details = session.query(HistoryModel).all()
        history_dicts = list()
        for history_detail in history_details:
            history_dict = history_detail.__dict__ 
            history_dict["category"] = categorize_date(history_dict["last_updated"])
            history_dicts.append(history_dict)
        history_dicts = sorted(history_dicts, key=lambda x: datetime.strptime(x["last_updated"], '%A, %B %d %Y %I:%M%p').strftime('%m/%d/%Y %H:%M:%S'), reverse=True)
        # print(history_dicts)
        return history_dicts


    @Slot(str)
    def remove_history_from_db(self, name):
        # if it has file in the trackdb folder remove it
        with Session() as session:
            history_to_delete = session.query(HistoryModel).filter_by(name=name).first()
            file_path = history_to_delete.address if history_to_delete.mode != "Single" else " " 
            if history_to_delete:
                session.delete(history_to_delete)
                session.commit()
                print(name, "removed")
        if os.path.exists(file_path):
            try:
                # Remove the file
                os.remove(file_path)
                print(f"File '{file_path}' has been removed.")
            except OSError as e:
                print(f"Error: {e}")
        else:
            print(f"File '{file_path}' does not exist.") 


    @Slot(str)
    def track_single(self, number:str) -> None:
        login = self.fetch_current_login()
        api_key = login['api_key']
        self.track_single_thread = QThread()
        self.track_single_worker = Worker( number=number, api_key=api_key, mode="Single")
        # a threader to collect ( worker: class, thread: Qthread, workerfn: function, args: [cor worker signal , cor mainFn ])
        threader(self.track_single_worker, self.track_single_thread, self.track_single_worker.track_single_work, [self.track_single_worker.single , self.track_single_main])
        

    def track_single_main(self, data:dict):
        # print(data)
        self.single.emit(data)


  


    @Slot(str)
    def load_excel(self, excel_name:str) -> None:
        self.load_excel_thread = QThread()
        self.load_excel_worker = Worker(excel_name=excel_name)
        # a threader to collect ( worker: class, thread: Qthread, workerfn: function, args: [cor worker signal , cor mainFn ])
        self.loaded_excel_worker, self.loaded_excel_thread  = threader(self.load_excel_worker, self.load_excel_thread, self.load_excel_worker.load_excel_work, [self.load_excel_worker.load , self.load_excel_main])
        
                

    def load_excel_main(self, data:dict):
        print(data)
        self.load.emit(data)


    @Slot(str)
    def change_sheet(self, sheetTitle):
        self.loaded_excel_worker.sheet = self.loaded_excel_worker.book[sheetTitle]
        max_row, max_col = get_max_row_and_column(self.loaded_excel_worker.sheet)
        print(self.loaded_excel_worker.sheet.title , max_row, max_col, "original", self.loaded_excel_worker.sheet.max_row, self.loaded_excel_worker.sheet.max_column)
        rows = [i for i in range(1, max_row + 1)]
        cols = [{'value': " ", 'width': 40 , "color":"#F9F9F9", "coordinate": "xy" }] + [{'value': column, 'width':self.loaded_excel_worker.sheet.column_dimensions[column].width*7 , "color":"#F9F9F9", "coordinate": "y"  } for column in col_format(max_col)]
        loaded = { "rows": json.dumps(rows) , "cols": json.dumps(cols) , "sheet_names": json.dumps(self.loaded_excel_worker.book.sheetnames) , "active_sheet_index": self.loaded_excel_worker.book.sheetnames.index(sheetTitle), "access": True } 
        self.loaded_excel_worker.update_sheet()
        self.load.emit(loaded)

    
    @Slot(result=list)
    def sheetList(self):
        self._sheetList = self.loaded_excel_worker.sheetList 
        print("excel sheet passed to gui")
        print(self._sheetList)
        return self._sheetList  
    

    @Slot(result=list)
    def getHighList(self):
        # print("list transfered")
        return self.loaded_excel_worker.highList 

    @Slot(result=dict)
    def getAllCellValues(self):
        # print("all cell transfered")
        return self.loaded_excel_worker.allCellValues

    @Slot(list,str, int)
    def track_excel(self, numbers:list, track_name, resume: int) -> None:
        numbers = [str(i) for i in numbers]
        login = self.fetch_current_login()
        api_key = login['api_key']
        self.track_excel_thread = QThread()
        self.track_excel_worker = Worker( api_key=api_key, numbers=numbers, resume=resume, track_name=track_name, mode="Excel")
        # a threader to collect ( worker: class, thread: Qthread, workerfn: function, args: [cor worker signal , cor mainFn ])
        self.excel_worker, self.excel_thread = threader(self.track_excel_worker, self.track_excel_thread, self.track_excel_worker.track_multiple_work, [self.track_excel_worker.multiple , self.track_excel_main])

    @Slot()
    def stop_excel(self):
        self.excel_worker.stop()    
        print("Excel stopped")


    def track_excel_main(self, data:dict):
        # print(data)
        self.excels.emit(data)



    @Slot(list,str, int)
    def track_paste(self, numbers:list, track_name, resume: int) -> None:
        numbers = [str(i) for i in numbers]
        login = self.fetch_current_login()
        api_key = login['api_key']
        self.track_paste_thread = QThread()
        self.track_paste_worker = Worker(api_key=api_key, numbers=numbers, resume=resume, track_name=track_name, mode="Paste")
        # a threader to collect ( worker: class, thread: Qthread, workerfn: function, args: [cor worker signal , cor mainFn ])
        self.paste_worker, self.paste_thread = threader(self.track_paste_worker, self.track_paste_thread, self.track_paste_worker.track_multiple_work, [self.track_paste_worker.multiple , self.track_paste_main])


    @Slot()
    def stop_paste(self):
        self.paste_worker.stop()    
        print("paste stopped")


    def track_paste_main(self, data:dict):
        # print(data)
        self.pastes.emit(data)    


   
    @Slot(str, str)
    def fetch_export(self, address, mode):
        self.fetch_export_thread = QThread()
        self.fetch_export_worker = Worker(address=address, mode=mode)
        # a threader to collect ( worker: class, thread: Qthread, workerfn: function, args: [cor worker signal , cor mainFn ])
        self.loaded_excel_worker, self.loaded_excel_thread  = threader(self.fetch_export_worker, self.fetch_export_thread, self.fetch_export_worker.fetch_export_work, [self.fetch_export_worker.fetch , self.fetch_export_main])
    


    def fetch_export_main(self, data:dict):
        # print(data)
        self.fetch.emit(data)



    @Slot(str, str, list)
    def save_fetched(self, path , extension,  headers):
        self.save_fetched_thread = QThread()
        self.save_fetched_worker = Worker(path=path , extension=extension, headers=headers)
        self.loaded_excel_worker, self.loaded_excel_thread  = threader(self.save_fetched_worker, self.save_fetched_thread, self.save_fetched_worker.save_fetched_work, [self.save_fetched_worker.save , self.save_fetched_main])

        # self.loaded_excel_worker.save_fetched() 
        
    def save_fetched_main(self, data:dict):
        self.save.emit(data)


    def fetch_current_login(self)-> dict:
        login = self.fetch_saved_logins()[self.fetch_current_login_index()]
        return login    
             

    def check_login_main(self,data:dict):
        api_key = data["api_key"]
        username = data["username"]
        if data["truth"]:
            with Session() as session:
                new_data = LoginModel(api_key=api_key, username=username  ,timestamp=str(datetime.now().strftime("%A, %B %d %Y %I:%M%p")))
                session.add(new_data)
                session.commit()
        self.check.emit([data["truth"], data["status"], api_key, username])



    @Slot(str,str)
    def login_to_db(self, api_key:str, username:str ) -> None:
        self.check_login_thread = QThread()
        # Step 3: Create a worker object
        self.check_login_worker = Worker(api_key=api_key, username=username)
        # a threader to collect ( worker: class, thread: Qthread, workerfn: function, args: [cor worker signal , cor mainFn ])
        threader(self.check_login_worker, self.check_login_thread, self.check_login_worker.check_login_work, [self.check_login_worker.check , self.check_login_main])

        
    @Slot(result=list)
    def get_saved_logins(self): #  has a column of just id, acctNo, username , password , timestamp
        return self.fetch_saved_logins()
        
    
    def fetch_saved_logins(self):
        with Session() as session:
            logins_details = session.query(LoginModel).all() 
            # Convert each row to a dictionary
            logins_dicts = [ filter_dict(logins_detail.__dict__, ['api_key', 'username']) for logins_detail in logins_details]
            print(logins_dicts)      
        return logins_dicts



    @Slot(str)
    def remove_login_from_db(self, username):
        with Session() as session:
            login_to_delete = session.query(LoginModel).filter_by(username=username).first()         
            if login_to_delete:
                # Delete the row
                session.delete(login_to_delete)
                session.commit()              
       
              
                    
    @Slot(int)
    def set_current_login_index(self, index):
        config = configparser.ConfigParser()
        config.read('settings.ini')
        config.set('Logins', 'currentLoginIndex', str(index))
        with open('settings.ini', 'w') as configfile:
            config.write(configfile)      


 
    @Slot(result=int)
    def get_current_login_index(self):
        return self.fetch_current_login_index()
         

    def fetch_current_login_index(self):
        config = configparser.ConfigParser()
        config.read('settings.ini')
        index = int(config.get('Logins', 'currentLoginIndex'))  
        return index  


    @Slot(str, str)
    def format_to_db(self, name: str , cols: str) -> None:
        with Session() as session:
            new_data = FormatModel(name=name, cols=cols)
            session.add(new_data)
            session.commit()


    @Slot(str)
    def remove_format_from_db(self, name):
        with Session() as session:
            format_to_delete = session.query(FormatModel).filter_by(name=name).first() 
            if format_to_delete:
                # Delete the row
                session.delete(format_to_delete)
                session.commit()  
        

    @Slot(result=list)
    def get_saved_formats(self):
        return self.fetch_saved_formats()


    def fetch_saved_formats(self):
        with Session() as session:
            format_details = session.query(FormatModel).all() 
            # Convert each row to a dictionary
            format_dicts =  [ filter_dict( format_detail.__dict__, ['cols', 'name']) for format_detail in format_details]            
            print(format_dicts)
        return format_dicts
    

    @Slot(str, result=list)
    def extract_awb_from_text(self, text_awbs):
        # print(GlogalVariable.pattern)
        list_awbs = re.findall(GlogalVariable.pattern2, text_awbs) # Note you will need this list in the future    
        return list_awbs        



    

        
    
        


   

if __name__ == "__main__":
    # Enable scaling
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    # Don't use resources, use local paths
    os.environ["QT_QUICK_CONTROLS_CONF"] = "qtquickcontrols2.conf"
    os.environ["QT_QUICK_CONTROLS_STYLE"] = "Material"
    QIcon.setThemeSearchPaths(['./icons'])
    app = QApplication(sys.argv)
    QIcon.setThemeName("gallery")
    engine = QQmlApplicationEngine()
    main = MainWindow()
    engine.rootContext().setContextProperty("Backend", main)
    engine.load(QUrl('index.qml'))
    sys.exit(app.exec())
    
    
